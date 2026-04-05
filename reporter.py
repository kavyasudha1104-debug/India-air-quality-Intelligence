import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

def get_fill(grade):
    grade = str(grade)
    if grade == "A":
        return PatternFill("solid", fgColor="00B050")
    elif grade == "B":
        return PatternFill("solid", fgColor="92D050")
    elif grade == "C":
        return PatternFill("solid", fgColor="FFFF00")
    elif grade == "D":
        return PatternFill("solid", fgColor="FF7700")
    else:
        return PatternFill("solid", fgColor="FF0000")

def generate_excel_report(scored_df, trends, alerts):
    wb = openpyxl.Workbook()

    # Sheet 1: City Rankings
    ws1 = wb.active
    ws1.title = "City Rankings"

    headers = ["Rank", "City", "AQI", "Category", "Health Score", "Grade"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B6")
        cell.alignment = Alignment(horizontal="center")

    for _, row in scored_df.iterrows():
        r = int(row["score_rank"]) + 1
        city_short = row["city"].split(",")[0]
        ws1.cell(row=r, column=1, value=int(row["score_rank"]))
        ws1.cell(row=r, column=2, value=city_short)
        ws1.cell(row=r, column=3, value=int(row["aqi"]))
        ws1.cell(row=r, column=4, value=str(row["category"]))
        ws1.cell(row=r, column=5, value=int(row["health_score"]))
        grade_cell = ws1.cell(row=r, column=6, value=str(row["grade"]))
        grade_cell.fill = get_fill(row["grade"])
        grade_cell.font = Font(bold=True)
        grade_cell.alignment = Alignment(horizontal="center")

    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["D"].width = 28
    for col in ["A", "C", "E", "F"]:
        ws1.column_dimensions[col].width = 14

    # Sheet 2: Alerts
    ws2 = wb.create_sheet("Alerts")
    ws2.cell(row=1, column=1, value="Alert Type").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor="2F75B6")
    ws2.cell(row=1, column=2, value="Message").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=1, column=2).fill = PatternFill("solid", fgColor="2F75B6")

    for i, alert in enumerate(alerts, 2):
        alert_type = alert.split(":")[0]
        ws2.cell(row=i, column=1, value=alert_type)
        ws2.cell(row=i, column=2, value=alert)
        if "CRITICAL" in alert:
            ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FF0000")
        elif "WARNING" in alert:
            ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FF7700")
        else:
            ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor="00B050")

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 60

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="India Air Quality Summary").font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value="Date").font = Font(bold=True)
    ws3.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d"))
    ws3.cell(row=4, column=1, value="Best City").font = Font(bold=True)
    ws3.cell(row=4, column=2, value=trends["best"].split(",")[0])
    ws3.cell(row=5, column=1, value="Worst City").font = Font(bold=True)
    ws3.cell(row=5, column=2, value=trends["worst"].split(",")[0])
    ws3.cell(row=6, column=1, value="Total Cities Tracked").font = Font(bold=True)
    ws3.cell(row=6, column=2, value=len(scored_df))
    ws3.cell(row=7, column=1, value="Critical Cities").font = Font(bold=True)
    ws3.cell(row=7, column=2, value=len(trends["critical"]))
    ws3.cell(row=8, column=1, value="Good Air Cities").font = Font(bold=True)
    ws3.cell(row=8, column=2, value=len(trends["good"]))
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 25

    filename = f"reports/AQI_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    print(f"\nExcel report saved: {filename}")